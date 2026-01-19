"""
Traffic Flow Visualization Generator

This module generates PNG plots of traffic flows in a sankey diagramm format from Excel data.
It supports general traffic and peak hour analysis (morning and afternoon).
"""

import io
import re
from typing import List, Tuple, Dict, Optional, Any   #Type hints
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg") # Agg is a non-interactive, off-screen rendering backend, Plots are rendered directly to image file
import matplotlib.pyplot as plt    
from matplotlib.patches import Polygon
from openpyxl import load_workbook
from datetime import datetime, timedelta


# --------------------- CONFIG ---------------------
# Minimum/maximum thickness for flow bands 
width_min = 0.1
width_max = 0.7

#PKW_Einheiten faktors
faktor_rad = 0.5
faktor_Linienbus = 1.5
faktor_lkwAnh = 2
faktor_sonst = 1.5

# 12 directions possible (R1..R12) -> endpoints (i,j)
DIR_TO_FLOW = {
    1:  (1, 24),
    2:  (2, 17),
    3:  (3, 10),
    4:  (6, 7),
    5:  (8, 23),
    6:  (16, 9),
    7:  (13, 12),
    8:  (14, 5),
    9:  (15, 22),
    10: (18, 19),
    11: (20, 11),
    12: (4, 21),
}

# Rectangles: define as unordered so orientation doesn't matter
RECT_FLOWS_U = {tuple(sorted(p)) for p in [(2, 17), (5, 14), (8, 23), (11, 20)]}

# Draw params
C = np.array([0.0, 0.0])    # center
R = 4.0                     # radius for placing points  
d = 1                       # distance from center line to middle point of group
inward = 0.9                # inward control for bezier curves (curvature strength)

FILL = "lightblue"
EDGE = "none"
EDGE_LW = 0.0

# Group slots
# Dict mapping (side, type) → list of port IDs
GROUP_SLOTS = {
    ("N", "dep"): [1, 2, 3],
    ("N", "arr"): [6, 5, 4],

    ("E", "dep"): [7, 8, 9],
    ("E", "arr"): [12, 11, 10],

    ("S", "dep"): [13, 14, 15],
    ("S", "arr"): [18, 17, 16],

    ("W", "dep"): [19, 20, 21],
    ("W", "arr"): [24, 23, 22],
}

# Dict side → Matplotlib color name.
SIDE_COLOR = {"N": "tab:blue", "E": "tab:orange", "S": "tab:green", "W": "tab:red"}

def add_flow_label_before_start(ax, A, side, text, color, fontsize=6):
    """Adds a the traffic near the start of the corresponding flow, before the flow begins"""
    A = np.asarray(A, float)        #Ensures A is a numpy float array.

    back = 0.15                     #How far outside the port the label sits

    #Locate start position and orientation based on side
    if side == "E":
        u_pos = np.array([1.0, 0.0])    
        angle_deg = 0
        ha, va = "left", "center"
    elif side == "W":
        u_pos = np.array([-1.0, 0.0])    
        angle_deg = 0
        ha, va = "right", "center"
    elif side == "N":
        u_pos = np.array([0.0, 1.0])     
        angle_deg = -90                  
        ha, va = "right", "center"
    else:  # "S"
        u_pos = np.array([0.0, -1.0])    
        angle_deg = -90
        ha, va = "left", "center"

    #Final label position
    pos = A + back * u_pos

    ax.text(
        pos[0], pos[1], text,
        rotation=angle_deg,
        rotation_mode="anchor",
        ha=ha, va=va,
        fontsize=fontsize,
        color=color,
        zorder=50,
        fontweight="bold",
        clip_on=False
    )

def add_side_span_line_and_total(ax,P,W,dep_ids,arr_ids,side,total_text,d_NS,d_WE,line_lw=3,line_color="black",text_color="black",text_fontsize=18,offset_line=0.9,offset_text=1.2,zorder=40):
    """
    Draw a span line between the two most outward edges of a side considering BOTH dep+arr ports.
    Outward edges are computed as coordinate +/- (W[pid]/2) along the variable axis.

    If dep or arr group is missing, synthesize the missing group by shifting the existing group
    by +/- 2*d along the variable axis (keeping same widths).

    Text is horizontal for N/S and vertical (top->bottom) for E/W.
    """

    # Determine variable axis + outward normal + dep<->arr shift
    if side in ("N", "S"):
        var_axis = 0  # x varies
        other_axis = 1
        nrm = np.array([0.0, +1.0]) if side == "N" else np.array([0.0, -1.0])
        shift = 2.0 * float(d_NS)
    else:
        var_axis = 1  # y varies
        other_axis = 0
        nrm = np.array([+1.0, 0.0]) if side == "E" else np.array([-1.0, 0.0])
        shift = 2.0 * float(d_WE)

    dep_ids = list(dep_ids) if dep_ids else []
    arr_ids = list(arr_ids) if arr_ids else []

    # Collect all real points (must have at least 2)
    real_pids = [pid for pid in (dep_ids + arr_ids) if pid in P and pid in W]
    if len(real_pids) < 2:
        return

    # We will build a list of "edge extents" along var_axis:
    # each item is (min_edge, max_edge, other_coord)
    extents = []

    def add_pid_extent(pid, delta_var=0.0):
        pt = np.array(P[pid], float).copy()
        pt[var_axis] += delta_var
        half = float(W[pid]) / 2.0
        min_edge = float(pt[var_axis] - half)
        max_edge = float(pt[var_axis] + half)
        extents.append((min_edge, max_edge, float(pt[other_axis])))

    # Add real extents
    for pid in real_pids:
        add_pid_extent(pid, delta_var=0.0)

    # If one group missing, synthesize by shifting the existing group's points
    if len(dep_ids) == 0 and len(arr_ids) > 0:
        for pid in arr_ids:
            if pid in P and pid in W:
                add_pid_extent(pid, delta_var=-shift)

    if len(arr_ids) == 0 and len(dep_ids) > 0:
        for pid in dep_ids:
            if pid in P and pid in W:
                add_pid_extent(pid, delta_var=+shift)

    if len(extents) < 2:
        return

    # Global min/max edges across all extents
    min_edge = min(e[0] for e in extents)
    max_edge = max(e[1] for e in extents)

    # Choose a stable coordinate on the other axis (average works well)
    other_mean = float(np.mean([e[2] for e in extents]))

    # Build line endpoints in data coords
    p1 = np.array([0.0, 0.0])
    p2 = np.array([0.0, 0.0])
    p1[var_axis] = min_edge
    p2[var_axis] = max_edge
    p1[other_axis] = other_mean
    p2[other_axis] = other_mean

    # Push line outward
    p1_line = p1 + offset_line * nrm
    p2_line = p2 + offset_line * nrm

    ax.plot(
        [p1_line[0], p2_line[0]],
        [p1_line[1], p2_line[1]],
        linewidth=line_lw,
        color=line_color,
        solid_capstyle="round",
        zorder=zorder,
        clip_on=False,
    )

    # Text position: midpoint, pushed further outward
    mid = 0.5 * (p1 + p2)
    pos_text = mid + offset_text * nrm

    # Rotation: E/W should read North -> South (top -> bottom)
    rotation = 270 if side in ("E", "W") else 0

    ax.text(
        pos_text[0], pos_text[1],
        str(total_text),
        ha="center", va="center",
        fontsize=text_fontsize,
        color=text_color,
        fontweight="bold",
        rotation=rotation,
        rotation_mode="anchor",
        zorder=zorder + 1,
        clip_on=False,
    )

def compute_side_sums(flows_present, kfz_array):
    """
    Compute sums per side for departing and arriving traffic.

    Returns:
      dep_kfz, arr_kfz, total_kfz
      and if bike_array provided: dep_bike, arr_bike, total_bike
    """

    # Map port id -> side for dep and arr ports
    dep_pid_to_side = {}
    arr_pid_to_side = {}
    for side in ("N", "E", "S", "W"):
        for pid in GROUP_SLOTS[(side, "dep")]:
            dep_pid_to_side[pid] = side
        for pid in GROUP_SLOTS[(side, "arr")]:
            arr_pid_to_side[pid] = side

    dep_kfz = {s: 0.0 for s in ("N", "E", "S", "W")}
    arr_kfz = {s: 0.0 for s in ("N", "E", "S", "W")}

    # --- KFZ sums ---
    for (i, j), kfz in zip(flows_present, kfz_array):
        if i in dep_pid_to_side and j in arr_pid_to_side:
            dep_side = dep_pid_to_side[i]
            arr_side = arr_pid_to_side[j]
        elif j in dep_pid_to_side and i in arr_pid_to_side:
            dep_side = dep_pid_to_side[j]
            arr_side = arr_pid_to_side[i]
        else:
            continue

        dep_kfz[dep_side] += float(kfz)
        arr_kfz[arr_side] += float(kfz)

    total_kfz = {s: dep_kfz[s] + arr_kfz[s] for s in ("N", "E", "S", "W")}
    
    return {
        "dep_kfz": dep_kfz,
        "arr_kfz": arr_kfz,
        "total_kfz": total_kfz,
    }

def calculate_width(direction_dic, tmin, tmax, gamma=1.0, PKW_Einheiten=False):
    """
    Calculate width array based on KFZ values using a GLOBAL mapping:
      - tmin -> width_min
      - tmax -> width_max
      - in between proportional (non-linear with gamma)

    gamma = 1.0  -> linear
    gamma < 1.0  -> more resolution for small flows (recommended: 0.5)
    gamma > 1.0  -> more resolution for large flows
    """
    if not PKW_Einheiten:
        traffic = np.array([sub_dic["kfz"] for sub_dic in direction_dic.values()], dtype=float)  #Extracts all kfz values from the dictionary (in iteration order)
    else:
        traffic = np.array([sub_dic["PKW_Total"] for sub_dic in direction_dic.values()], dtype=float)  #Extracts all PKW_Einheiten values from the dictionary (in iteration order)

    if traffic.size == 1 or np.isclose(tmax, tmin):     #If only one flow or all flows equal (no scale), give them all mid-width
        return np.round(np.full_like(traffic, (width_min + width_max) / 2.0), 2)

    norm = (traffic - tmin) / (tmax - tmin)     #Normalize to [0,1]
    norm = np.clip(norm, 0.0, 1.0)      #Ensure norm within [0,1]

    widths = width_min + (norm ** gamma) * (width_max - width_min)      #Scale to [width_min, width_max] using gamma correction
    return np.round(widths, 2)      #Returns all widths 

def build_direction_dic(sheets, peak_idx):
    """Build direction dictionary for a given peak index from sheets starting with R."""
    dic = {}
    for sheet_name, df in sheets.items():       #Iterate over all sheets
        if sheet_name.startswith("R"):          #Only process sheets starting with "R"
            kfz_sum = df.iloc[peak_idx:peak_idx+4, 2:9].sum().sum()     #Sum kfz values in the specified range (4 rows starting at peak_idx, columns 2 to 8)
            total_sum = df.iloc[peak_idx:peak_idx+4, 1:9].sum().sum()   #Sum total values (kfz + rad) in the specified range (4 rows starting at peak_idx, columns 1 to 8)
            SV_sum = df.iloc[peak_idx:peak_idx+4, 4:9].sum().sum()
            dic[sheet_name] = {
                "total": total_sum,
                "kfz": kfz_sum,
                "rad": total_sum - kfz_sum,
                "Summe_SV": SV_sum
            }
    return dic

def PKW_Einheiten_traffic_dic(sheets, peak_idx):
    dic = {}
    for sheet_name, df in sheets.items():
        if sheet_name.startswith("R"):
            rad = df.iloc[peak_idx:peak_idx+4, 1].sum() * faktor_rad
            einsp = df.iloc[peak_idx:peak_idx+4, 2].sum()
            PKW = df.iloc[peak_idx:peak_idx+4, 3].sum()
            Linienbus = df.iloc[peak_idx:peak_idx+4, 4].sum() * faktor_Linienbus
            Reisebus = df.iloc[peak_idx:peak_idx+4, 5].sum() * faktor_Linienbus
            LKW = df.iloc[peak_idx:peak_idx+4, 6].sum() * faktor_Linienbus
            LKW_Anh = df.iloc[peak_idx:peak_idx+4, 7].sum() * faktor_lkwAnh
            sons = df.iloc[peak_idx:peak_idx+4, 8].sum() * faktor_sonst
            dic[sheet_name] = {
                "PKW_Total": round(rad + einsp + PKW + Linienbus + Reisebus + LKW + LKW_Anh + sons),
                "Summe_SV": round(Linienbus + Reisebus + LKW + LKW_Anh + sons)
            }
    return dic

def _sv_stats(total: float, sv: float) -> Dict[str, float]:
    total = float(total)
    sv = float(sv)
    share = (sv / total * 100.0) if total > 0 else 0.0
    return {"total": total, "sv": sv, "sv_share_pct": round(share, 2)}

# --------------------- GEOMETRY HELPERS ---------------------
def segment_rectangle(A, B, width):
    """Create a rectangular segment between points A and B with given width."""
    A = np.asarray(A, float)
    B = np.asarray(B, float)
    v = B - A
    L = np.hypot(v[0], v[1])
    u = v / L
    n = np.array([-u[1], u[0]])
    off = (width / 2.0) * n
    return np.vstack([A + off, B + off, B - off, A - off])

def inward_ctrl(Z, P, inward_amount):
    """Calculate inward control point for bezier curves."""
    return P + inward_amount * (Z - P)

def bezier_points(P0, P1, P2, P3, n=250):
    """Generate points along a bezier curve."""
    t = np.linspace(0, 1, n)[:, None]
    return ((1-t)**3)*P0 + 3*((1-t)**2)*t*P1 + 3*(1-t)*(t**2)*P2 + (t**3)*P3

def bezier_tangent(P0, P1, P2, P3, n=250):
    """Calculate tangents along a bezier curve."""
    t = np.linspace(0, 1, n)[:, None]
    return (3*((1-t)**2)*(P1-P0) + 6*(1-t)*t*(P2-P1) + 3*(t**2)*(P3-P2))

def bezier_ribbon_polygon(P0, P1, P2, P3, width, n=250, eps=1e-12):
    """Create a ribbon polygon along a bezier curve."""
    pts = bezier_points(P0, P1, P2, P3, n)
    tan = bezier_tangent(P0, P1, P2, P3, n)

    L = np.maximum(np.hypot(tan[:, 0], tan[:, 1]), eps)
    u = tan / L[:, None]
    nrm = np.column_stack([-u[:, 1], u[:, 0]])

    off = (width / 2.0) * nrm
    left = pts + off
    right = pts - off
    return np.vstack([left, right[::-1]])

def add_patch(ax, poly, color=None):
    """Add a polygon patch to the axes."""
    ax.add_patch(
        Polygon(
            poly, closed=True,
            facecolor=color if color is not None else FILL,
            edgecolor=EDGE, linewidth=EDGE_LW
        )
    )

def add_bezier_ribbon(ax, A, B, Z, width, color):
    """Add a bezier ribbon between A and B via Z."""
    P0, P3 = A, B
    P1 = inward_ctrl(Z, A, inward)
    P2 = inward_ctrl(Z, B, inward)
    poly = bezier_ribbon_polygon(P0, P1, P2, P3, width=width)
    add_patch(ax, poly, color)

def place_group_variable(P, fixed_axis, fixed_val, ids, mid_val, dir_to_axis, W):
    """Place points for a group of slots.
    - fixed axis: 0 for x fixed, 1 for y fixed
    - fixed val: value on fixed axis 
        (N: fixed axis = 1, fixed_val = +R, 
         S: fixed axis = 1, fixed_val = -R,
         E: fixed axis = 0, fixed_val = +R,
         W: fixed axis = 0, fixed_val = -R)
    - ids: list of point IDs ([1,2,3], etc.)
    - mid_val: midpoint value on the variable axis (+d or -d)
    - dir_to_axis: direction to the center axis (1 or -1)
    - W: width dictionary
    """
    if not ids: #if group empty, exit 
        return

    widths = [W[i] for i in ids]      #Get widths for all points in the group, order matter, determines left-to-right placement
    span = sum(widths)                #Total span of the group (sum of widths)

    offsets = []
    acc = -span / 2.0                  #Start at negative half-span, running currently accumulated offset
    for w in widths:
        offsets.append(acc + w / 2.0)
        acc += w

    #Point placement
    for pid, off in zip(ids, offsets):
        pt = np.array([0.0, 0.0], float)
        pt[fixed_axis] = fixed_val
        pt[1 - fixed_axis] = mid_val + dir_to_axis * off     #Position on variable axis, works such that the center of the group lays at mid_val from center line
        P[pid] = C + pt

def align_rect_pairs_shift_groups(P: Dict[int, np.ndarray], pairs: List[Tuple[int, int]]) -> None:
    """
    Post-process already-computed port positions P so that the rectangle pairs
    (2,17), (5,14), (8,23), (11,20) are aligned.

    Alignment rule (per pair):
      - Use the midpoint of the pair's *variable* coordinate
      - Apply the required delta to the ENTIRE group (side, dep/arr) of each endpoint,
        so neighboring ports move together and flows still start/end flush.

    Variable axis:
      - N/S groups vary in x  -> axis 0
      - E/W groups vary in y  -> axis 1

    Requires GROUP_SLOTS to be defined (as in your module).
    """
    # --- local helpers (kept inside this single function) ---
    def _pid_to_group_key(pid):
        for side in ("N", "E", "S", "W"):
            if pid in GROUP_SLOTS[(side, "dep")]:
                return (side, "dep")
            if pid in GROUP_SLOTS[(side, "arr")]:
                return (side, "arr")
        return None

    def _var_axis_from_side(side):
        return 0 if side in ("N", "S") else 1

    # --- apply constraints ---
    for a, b in pairs:
        if a not in P or b not in P:
            continue

        ga = _pid_to_group_key(a)
        gb = _pid_to_group_key(b)
        if ga is None or gb is None:
            continue

        side_a, _ = ga
        side_b, _ = gb
        ax_a = _var_axis_from_side(side_a)
        ax_b = _var_axis_from_side(side_b)
        if ax_a != ax_b:
            # safety: don't try to align across different variable axes
            continue
        var_axis = ax_a

        Pa = np.array(P[a], float)
        Pb = np.array(P[b], float)
        mid = 0.5 * (Pa[var_axis] + Pb[var_axis])

        # shift full group containing a
        delta_a = mid - Pa[var_axis]
        for pid in GROUP_SLOTS[ga]:
            if pid in P:
                Ppid = np.array(P[pid], float)
                Ppid[var_axis] += delta_a
                P[pid] = Ppid

        # shift full group containing b
        delta_b = mid - Pb[var_axis]
        for pid in GROUP_SLOTS[gb]:
            if pid in P:
                Ppid = np.array(P[pid], float)
                Ppid[var_axis] += delta_b
                P[pid] = Ppid

    """
    Force each (a,b) pair to share the same 'variable' coordinate by setting
    both to the midpoint of their current variable coordinate.

    Variable axis:
      - If the points are on N/S (y is +/-R), variable axis is x (axis 0)
      - If the points are on E/W (x is +/-R), variable axis is y (axis 1)

    This runs AFTER P has been computed.
    """
    for a, b in pairs:
        if a not in P or b not in P:
            continue

        Pa = np.array(P[a], float)
        Pb = np.array(P[b], float)

        # Decide if this pair is N/S-like or E/W-like based on which coordinate is "fixed"
        # N/S points have y near +/-R (so y has large abs), E/W points have x near +/-R.
        if abs(Pa[1]) >= abs(Pa[0]) and abs(Pb[1]) >= abs(Pb[0]):
            var_axis = 0  # align x
        else:
            var_axis = 1  # align y

        mid = 0.5 * (Pa[var_axis] + Pb[var_axis])

        Pa[var_axis] = mid
        Pb[var_axis] = mid

        P[a] = Pa
        P[b] = Pb

def add_group_arrow(ax, P, W, group_ids, side, outward=True, color="k", zorder=10,
                    label: Optional[str] = None, label_color: str = "white",
                    label_fontsize: int = 9):
    """Add an arrow for a group of slots, optionally with a text label inside."""
    ids = list(group_ids)
    pts = np.array([P[i] for i in ids], float)

    if side in ("N", "S"):
        var_axis = 0  # x
        nrm = np.array([0.0, +1.0]) if side == "N" else np.array([0.0, -1.0])
    else:
        var_axis = 1  # y
        nrm = np.array([+1.0, 0.0]) if side == "E" else np.array([-1.0, 0.0])

    var = pts[:, var_axis]
    far_idx = int(np.argmax(np.abs(var)))
    clo_idx = int(np.argmin(np.abs(var)))

    pid_far = ids[far_idx]
    pid_clo = ids[clo_idx]

    P_far = np.array(P[pid_far], float)
    P_clo = np.array(P[pid_clo], float)

    s_far = np.sign(P_far[var_axis]) or 1.0
    d_far = (W[pid_far] / 2.0) * s_far
    d_clo = -(W[pid_clo] / 2.0) * s_far

    base_far = P_far.copy(); base_far[var_axis] += d_far
    base_clo = P_clo.copy(); base_clo[var_axis] += d_clo

    base_center = 0.5 * (base_far + base_clo)
    tip = base_center + (nrm * 0.5 if outward else -nrm * 0.5)

    tri = np.vstack([tip, base_far, base_clo])

    ax.add_patch(Polygon(tri, closed=True, facecolor=color, edgecolor="none", zorder=zorder))

    # ---- Label inside arrow ----
    if label is not None:
        centroid = tri.mean(axis=0)
        ax.text(
            centroid[0], centroid[1], label,
            ha="center", va="center",
            fontsize=label_fontsize,
            color=label_color,
            zorder=zorder + 1,
            fontweight="bold",
            clip_on=False
        )

def create_plot(kfz, bike, width, flows_present, verkehrszählungsort, suffix, start_time, end_time, side_colors, d_NS, d_WE):
    """Create a PNG plot for given traffic and width data.
    kfz: numpy array of KFZ flow magnitudes aligned with flows_present
    bike: numpy array of bicycle flow magnitudes aligned with flows_present
    width: numpy array of ribbon widths aligned with flows_present
    flows_present: list of (i,j) present
    present_dirs: list of sheet names ["R1","R2",...] (mostly informational)
    verkehrszählungsort: location name from Excel
    suffix: string for filename (full_day, morning_peak, ...)
    side_colors: optional dict overriding SIDE_COLOR
    """
    # Update SIDE_COLOR with user-provided side_colors
    if side_colors:
        SIDE_COLOR.update(side_colors)

    # Width mapping already done
    flow_width = {(i, j): w for (i, j), w in zip(flows_present, width)}
    
    #Example: flows_present = [(1,24),(2,17),(3,10),(6,7)]
    #         width         = [0.65,    0.30,   0.15,  0.55]
    #         flow_width = {(1,24): 0.65, (2,17): 0.30, (3,10): 0.15, (6,7) : 0.55}

    # Assigns a width to each port ID
    W = {}
    for (i, j), w in flow_width.items():
        W[i], W[j] = w, w
    active_points = set(W.keys())

    # Map each flow (i,j) -> traffic value (same ordering as flows_present), only for the text of the traffic
    flow_kfz  = {(i, j): float(v) for (i, j), v in zip(flows_present, kfz)}
    flow_bike = {(i, j): float(v) for (i, j), v in zip(flows_present, bike)}
    show_departure_labels = True

    # Active groups, only include points that are active
    GROUP_ACTIVE = {
        key: [pid for pid in values if pid in active_points]
        for key, values in GROUP_SLOTS.items()
    }

    # Colors
    departing_points = set()  
    pid_to_side = {}      
    point_to_color = {}
    for side in ("N", "E", "S", "W"):
        for p in GROUP_ACTIVE[(side, "dep")]:
            departing_points.add(p)                 #Add departing point to set
            pid_to_side[p] = side                   #Map port ID to side
            point_to_color[p] = SIDE_COLOR[side]    #Map port ID to color

    def flow_color(i, j, default="lightblue"):      #Determine flow color based on departing points
        if i in departing_points:
            return point_to_color[i]
        if j in departing_points:
            return point_to_color[j]
        return default

    # Place points
    P = {}
    place_group_variable(P, 1, +R, GROUP_ACTIVE[("N","dep")], -d_NS, +1, W)
    place_group_variable(P, 1, +R, GROUP_ACTIVE[("N","arr")], +d_NS, -1, W)

    place_group_variable(P, 0, +R, GROUP_ACTIVE[("E","dep")], +d_WE, -1, W)
    place_group_variable(P, 0, +R, GROUP_ACTIVE[("E","arr")], -d_WE, +1, W)

    place_group_variable(P, 1, -R, GROUP_ACTIVE[("S","dep")], +d_NS, -1, W)
    place_group_variable(P, 1, -R, GROUP_ACTIVE[("S","arr")], -d_NS, +1, W)

    place_group_variable(P, 0, -R, GROUP_ACTIVE[("W","dep")], -d_WE, +1, W)
    place_group_variable(P, 0, -R, GROUP_ACTIVE[("W","arr")], +d_WE, -1, W)

    # --- ALIGN RECT PAIRS (post-placement) ---
    align_rect_pairs_shift_groups(
        P,
        pairs=[(2, 17), (5, 14), (8, 23), (11, 20)]
    )
    
    
    # Plot
    fig, ax = plt.subplots(figsize=(10, 10))  #fig is the whole image canvas, ax is the coordinate system where shapes are drawn

    for (i, j) in flows_present:
        if i not in P or j not in P:
            continue

        A, B = P[i], P[j]
        w = flow_width[(i, j)]
        col = flow_color(i, j)

        if tuple(sorted((i, j))) in RECT_FLOWS_U:
            add_patch(ax, segment_rectangle(A, B, w), col)
        else:
            Z = C + np.array([A[0], B[1]])
            add_bezier_ribbon(ax, A, B, Z, w, col)
            
        # ---------- LABEL BEFORE START ----------
        if show_departure_labels:
            start_pid = None

            # only label flows that start at a departing point
            if i in departing_points:
                start_pid = i
            elif j in departing_points:
                start_pid = j

            if start_pid is not None:
                Astart = np.asarray(P[start_pid], float)
                side = pid_to_side[start_pid]
                kfz_val = flow_kfz[(i, j)]
                bike_val = flow_bike[(i, j)]
                txt = f"{int(round(kfz_val))} | {int(round(bike_val))}"
                add_flow_label_before_start(ax, Astart, side, txt, color=col, fontsize=6)

    # ---------- GROUP ARROWS ----------
    side_sums = compute_side_sums(flows_present, kfz)
    dep_kfz_by_side = side_sums["dep_kfz"]
    arr_kfz_by_side = side_sums["arr_kfz"]
    total_kfz_by_side = side_sums["total_kfz"]
    for side in ("N", "E", "S", "W"):
        ids_dep = GROUP_ACTIVE[(side, "dep")]
        if len(ids_dep) >= 2:
            dep_label = str(int(round(dep_kfz_by_side.get(side, 0.0))))
            add_group_arrow(
                ax, P, W, ids_dep, side,
                outward=False, color="k",
                label=dep_label, label_color="white", label_fontsize=6
            )

        ids_arr = GROUP_ACTIVE[(side, "arr")]
        if len(ids_arr) >= 2:
            arr_label = str(int(round(arr_kfz_by_side.get(side, 0.0))))
            add_group_arrow(
                ax, P, W, ids_arr, side,
                outward=True, color="k",
                label=arr_label, label_color="white", label_fontsize=6
            )
            total_val = int(round(total_kfz_by_side.get(side, 0.0)))
            add_side_span_line_and_total(
                    ax,
                    P,
                    W,
                    dep_ids=ids_dep,
                    arr_ids=ids_arr,
                    side=side,
                    total_text=total_val,
                    d_NS=d_NS,
                    d_WE=d_WE,
                    line_lw=3,
                    text_fontsize=18,
                    offset_line=0.9,
                    offset_text=1.25,
                )

    ax.set_aspect("equal", adjustable="box")
    pad = 1.4
    ax.set_xlim(-R - pad, R + pad)
    ax.set_ylim(-R - pad, R + pad)
    ax.set_axis_off()

    # Return PNG bytes (no filesystem)
    buf = io.BytesIO()                  #raw bytes like a file (so you can read Excel and write PNGs without saving to disk)
    fig.savefig(buf, format="png", transparent=True, bbox_inches="tight", dpi=300)
    plt.close(fig)

    safe_name = re.sub(r"[^\w\-]+", "_", str(verkehrszählungsort))
    filename = f"VZ_{safe_name}_{suffix}_{start_time}_{end_time}.png"
    return buf.getvalue(), filename

# --------------------- MAIN GENERATOR ---------------------
def generate_png_from_excel(excel_bytes: bytes, side_colors: Optional[Dict[str, str]] = None, d_NS: float = 1, d_WE: float = 1, mode: str = "KFZ", use_custom_window: bool = False, custom_start_time: Optional[str] = None) -> Tuple[List[Tuple[bytes, str]], Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

    ws_deckblatt = wb["Deckbl."]
    verkehrszählungsort = ws_deckblatt["C8"].value

    # Load sheets for peak calculation
    sheets = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, header=None)
    
    first_R_df = None 
    for sheet_name, df in sheets.items():
        if sheet_name.startswith("R"):
                if first_R_df is None:
                    first_R_df = df
                    
    summe_idx = None
    if first_R_df is None:
            raise ValueError("No R sheets found – first_R_df was never assigned")
    for i, val in enumerate(first_R_df.iloc[:, 0]):
        if isinstance(val, str) and "SUMME" in val.upper():
            summe_idx = i
            break
    
    if summe_idx is None:
        raise ValueError("SUMME row not found")
    
    summe_row_number = summe_idx+1

    def _parse_interval(cell_value: Any) -> tuple[Optional[str], Optional[str]]:
        """
        Parse a time cell like '07:00-07:15' (also handles spaces and en-dash).
        Returns ('07:00','07:15') or (None,None) if not parseable.
        """
        if cell_value is None:
            return None, None
        s = str(cell_value).strip()
        s = s.replace("–", "-").replace("—", "-")
        s = s.replace(" ", "")
        if "-" not in s:
            return None, None
        a, b = s.split("-", 1)
        if len(a) == 5 and len(b) == 5:
            return a, b
        return None, None


    def _find_row_for_start(hhmm: str) -> int:
        """
        Find the row where the time window starts at hhmm.
        Excel format is like '7:45-8:00'.
        """
        for i in range(13, summe_idx):
            cell = str(first_R_df.iloc[i, 0])
            start = cell.split("-")[0].strip()

            # pad hour so '7:45' -> '07:45'
            h, m = start.split(":")
            start_norm = f"{int(h):02d}:{m}"

            if start_norm == hhmm:
                return i

        raise ValueError(f"Custom start time {hhmm} not found in Excel.")

    # Read directions
    direction_dic = {}
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("R"):
            ws = wb[sheet_name]
            total = ws[f"B{summe_row_number}"].value + ws[f"C{summe_row_number}"].value + ws[f"D{summe_row_number}"].value + ws[f"E{summe_row_number}"].value + ws[f"F{summe_row_number}"].value + ws[f"G{summe_row_number}"].value  + ws[f"H{summe_row_number}"].value  + ws[f"I{summe_row_number}"].value
            direction_dic[sheet_name] = {
                "total": total,
                "kfz": total - ws[f"B{summe_row_number}"].value,
                "rad": ws[f"B{summe_row_number}"].value,
                "Summe_SV": ws[f"E{summe_row_number}"].value + ws[f"F{summe_row_number}"].value + ws[f"G{summe_row_number}"].value  + ws[f"H{summe_row_number}"].value  + ws[f"I{summe_row_number}"].value
            }
    
    #PKW Einheiten
    PKW_direction_general_dic = {} 
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("R"):
            ws = wb[sheet_name]
            rad = ws[f"B{summe_row_number}"].value * faktor_rad
            einsp = ws[f"C{summe_row_number}"].value
            PKW = ws[f"D{summe_row_number}"].value
            Linienbus = ws[f"E{summe_row_number}"].value * faktor_Linienbus
            Reisebus = ws[f"F{summe_row_number}"].value * faktor_Linienbus
            LKW = ws[f"G{summe_row_number}"].value * faktor_Linienbus
            LKW_Anh = ws[f"H{summe_row_number}"].value * faktor_lkwAnh
            sons = ws[f"I{summe_row_number}"].value * faktor_sonst
            PKW_direction_general_dic[sheet_name] = {
                "PKW_Total": round(rad + einsp + PKW + Linienbus + Reisebus + LKW + LKW_Anh + sons),
                "Summe_SV": round(Linienbus + Reisebus + LKW + LKW_Anh + sons)
            }
    
    # Find peaks
    kfz_morning_peak = 0
    kfz_afternoon_peak = 0
    for idx in range(13, summe_idx-3):  # sliding window of 4 rows
        kfz_block_sum = 0
        
        for sheet_name, df in sheets.items():
            if sheet_name.startswith("R"):
                kfz_sheet_block_sum = df.iloc[idx:idx+4, 2:9].sum().sum()
                kfz_block_sum += kfz_sheet_block_sum
                
        # read time from first_R_df (stable reference)
        if first_R_df is None:
            raise ValueError("No R sheets found – first_R_df was never assigned")
        time_start = first_R_df.iloc[idx, 0]
        time_end   = first_R_df.iloc[idx+3, 0]

        if idx < summe_idx/2 and kfz_block_sum > kfz_morning_peak:
            kfz_morning_peak = kfz_block_sum
            morning_time_start = time_start
            morning_time_end = time_end
            morning_start_idx = idx

        if idx >= summe_idx/2 and kfz_block_sum > kfz_afternoon_peak:
            kfz_afternoon_peak = kfz_block_sum
            afternoon_time_start = time_start
            afternoon_time_end = time_end
            afternoon_peak_start_idx = idx
    
    col = 1

    first_idx = None
    last_idx = None
    started = False
    
    if first_R_df is None:
        raise ValueError("No R sheets found – first_R_df was never assigned")

    for row_idx in range(13, summe_idx):
        current_value = first_R_df.iloc[row_idx, col]

        if pd.notna(current_value) and not started:
            first_idx = row_idx
            last_idx = row_idx
            started = True
        elif pd.notna(current_value) and started:
            last_idx = row_idx
        elif pd.isna(current_value) and started:
            break
        
    if first_idx is None or last_idx is None:
        raise ValueError("No non-NaN data found in the specified range (rows 14–81)")
    day_start_time = str(first_R_df.iloc[first_idx, 0]).split("-")[0]
    day_end_time   = str(first_R_df.iloc[last_idx, 0]).split("-")[-1]
    morning_time_start = str(morning_time_start).split("-")[0]
    morning_time_end   = str(morning_time_end).split("-")[-1]

    afternoon_time_start = str(afternoon_time_start).split("-")[0]
    afternoon_time_end   = str(afternoon_time_end).split("-")[-1]
    
    #KFZ
    kfz_Tag_Summe = sum(value["kfz"] for value in direction_dic.values())
    kfz_Tag_SV = sum(value["Summe_SV"] for value in direction_dic.values())
    
    direction_morning_dic = build_direction_dic(sheets, morning_start_idx)
    direction_afternoon_dic = build_direction_dic(sheets, afternoon_peak_start_idx)
    
    direction_custom_dic = None
    PKW_Einheiten_traffic_custom = None
    custom_time_start = None
    custom_time_end = None
    custom_start_idx = None

    if use_custom_window:
        if not custom_start_time:
            raise ValueError("use_custom_window=True but custom_start_time is None")

        # end = start + 1 hour
        start_dt = datetime.strptime(custom_start_time, "%H:%M")
        custom_time_start = custom_start_time
        custom_time_end = (start_dt + timedelta(hours=1)).strftime("%H:%M")

        # locate the row where the interval starts at custom_time_start
        custom_start_idx = _find_row_for_start(custom_time_start)

        # we assume 15-min steps => 1 hour = 4 rows
        # make sure we don't go beyond SUMME
        if custom_start_idx + 3 >= summe_idx:
            raise ValueError("Custom 1h window exceeds available data in Excel.")

        direction_custom_dic = build_direction_dic(sheets, custom_start_idx)
        PKW_Einheiten_traffic_custom = PKW_Einheiten_traffic_dic(sheets, custom_start_idx)
    
    kfz_morning_summe = sum(value["kfz"] for value in direction_morning_dic.values())
    kfz_afternoon_summe = sum(value["kfz"] for value in direction_afternoon_dic.values())
    kfz_SV_morning = sum(value["Summe_SV"] for value in direction_morning_dic.values())
    kfz_SV_afternoon = sum(value["Summe_SV"] for value in direction_afternoon_dic.values())     
    
    #PKV Einheiten 
    PKW_Einheiten_Tag_Summe = sum(value["PKW_Total"] for value in PKW_direction_general_dic.values())
    PKW_Einheiten_Tag_SV = sum(value["Summe_SV"] for value in PKW_direction_general_dic.values())

    
    PKW_Einheiten_traffic_morning = PKW_Einheiten_traffic_dic(sheets, morning_start_idx)
    PKW_Einheiten_traffic_afternoon = PKW_Einheiten_traffic_dic(sheets, afternoon_peak_start_idx)

    PKW_Einheiten_morning_summe = sum(value["PKW_Total"] for value in PKW_Einheiten_traffic_morning.values())
    PKW_Einheiten_afternoon_summe = sum(value["PKW_Total"] for value in PKW_Einheiten_traffic_afternoon.values())

    PKW_Einheiten_SV_morning = sum(value["Summe_SV"] for value in PKW_Einheiten_traffic_morning.values())
    PKW_Einheiten_SV_afternoon = sum(value["Summe_SV"] for value in PKW_Einheiten_traffic_afternoon.values())

    kfz_sv_full = _sv_stats(kfz_Tag_Summe, kfz_Tag_SV)
    kfz_sv_morning = _sv_stats(kfz_morning_summe, kfz_SV_morning)
    kfz_sv_afternoon = _sv_stats(kfz_afternoon_summe, kfz_SV_afternoon)

    pkw_sv_full = _sv_stats(PKW_Einheiten_Tag_Summe, PKW_Einheiten_Tag_SV)
    pkw_sv_morning = _sv_stats(PKW_Einheiten_morning_summe, PKW_Einheiten_SV_morning)
    pkw_sv_afternoon = _sv_stats(PKW_Einheiten_afternoon_summe, PKW_Einheiten_SV_afternoon)

    # --- Ensure consistent ordering (important!) ---
    present_dirnums = sorted(int(name[1:]) for name in direction_dic.keys()) #Convert "R1" → 1, etc, and sorts
    if not present_dirnums:
        raise ValueError("No 'R*' sheets found. Nothing to plot.")

    present_dirs = [f"R{k}" for k in present_dirnums]       #ordered list of sheet names
    flows_present = [DIR_TO_FLOW[k] for k in present_dirnums]       #ordered list of edges (i,j) matching those directions

    # Reorder dictionaries so their .values() match present_dirs order
    direction_dic = {k: direction_dic[k] for k in present_dirs}
    direction_morning_dic = {k: direction_morning_dic[k] for k in present_dirs}
    direction_afternoon_dic = {k: direction_afternoon_dic[k] for k in present_dirs}

    # --- Global min/max across ALL three datasets ---
    all_kfz = []
    for dir_ in (direction_dic, direction_morning_dic, direction_afternoon_dic):
        all_kfz.extend(v["kfz"] for v in dir_.values())

    if direction_custom_dic is not None:
        all_kfz.extend(v["kfz"] for v in direction_custom_dic.values())

    tmin = float(min(all_kfz))
    tmax = float(max(all_kfz))

    # --- Calculate widths on shared scale ---
    gamma = 0.5  # <--- more resolution; set to 1.0 for strict linear

    width_general = calculate_width(direction_dic, tmin, tmax, gamma=gamma, PKW_Einheiten=False)
    width_morning_peak = calculate_width(direction_morning_dic, tmin, tmax, gamma=gamma, PKW_Einheiten=False)
    width_afternoon_peak = calculate_width(direction_afternoon_dic, tmin, tmax, gamma=gamma, PKW_Einheiten=False)
    width_PKW_general = calculate_width(PKW_direction_general_dic, tmin, tmax, gamma=gamma, PKW_Einheiten=True)
    width_PKW_morning = calculate_width(PKW_Einheiten_traffic_morning, tmin, tmax, gamma=gamma, PKW_Einheiten=True)
    width_PKW_afternoon = calculate_width(PKW_Einheiten_traffic_afternoon, tmin, tmax, gamma=gamma, PKW_Einheiten=True)
    
    width_custom = None
    width_PKW_custom = None
    
    kfz_custom = None
    bike_custom = None
    PKW_custom = None

    side_custom = None
    PKW_side_custom = None
        
    kfz_sv_custom = None
    pkw_sv_custom = None

    if direction_custom_dic is not None and PKW_Einheiten_traffic_custom is not None:
        width_custom = calculate_width(direction_custom_dic, tmin, tmax, gamma=gamma, PKW_Einheiten=False)
        width_PKW_custom = calculate_width(PKW_Einheiten_traffic_custom, tmin, tmax, gamma=gamma, PKW_Einheiten=True)
        
        # reorder to present_dirs order (important!)
        direction_custom_dic = {k: direction_custom_dic[k] for k in present_dirs}
        PKW_Einheiten_traffic_custom = {k: PKW_Einheiten_traffic_custom[k] for k in present_dirs}

        kfz_custom = np.array([direction_custom_dic[name]["kfz"] for name in present_dirs], dtype=float)
        bike_custom = np.array([direction_custom_dic[name]["rad"] for name in present_dirs], dtype=float)
        PKW_custom = np.array([PKW_Einheiten_traffic_custom[name]["PKW_Total"] for name in present_dirs], dtype=float)

        side_custom = compute_side_sums(flows_present, kfz_custom)
        PKW_side_custom = compute_side_sums(flows_present, PKW_custom)
        
        kfz_custom_sum = sum(value["kfz"] for value in direction_custom_dic.values())
        kfz_custom_sv  = sum(value["Summe_SV"] for value in direction_custom_dic.values())
        kfz_sv_custom = _sv_stats(kfz_custom_sum, kfz_custom_sv)

        pkw_custom_sum = sum(value["PKW_Total"] for value in PKW_Einheiten_traffic_custom.values())
        pkw_custom_sv  = sum(value["Summe_SV"] for value in PKW_Einheiten_traffic_custom.values())
        pkw_sv_custom = _sv_stats(pkw_custom_sum, pkw_custom_sv)   


    present_dirnums = sorted(int(name[1:]) for name in direction_dic.keys())
    if not present_dirnums:
        raise ValueError("No 'R*' sheets found. Nothing to plot.")

    kfz_general = np.array([direction_dic[name]["kfz"] for name in present_dirs], dtype=float)
    kfz_morning = np.array([direction_morning_dic[name]["kfz"] for name in present_dirs], dtype=float)
    kfz_afternoon = np.array([direction_afternoon_dic[name]["kfz"] for name in present_dirs], dtype=float)

    bike_general = np.array([direction_dic[name]["rad"] for name in present_dirs], dtype=float)
    bike_morning = np.array([direction_morning_dic[name]["rad"] for name in present_dirs], dtype=float)
    bike_afternoon = np.array([direction_afternoon_dic[name]["rad"] for name in present_dirs], dtype=float)

    PKW_general = np.array([PKW_direction_general_dic[name]["PKW_Total"] for name in present_dirs], dtype=float)
    PKW_morning = np.array([PKW_Einheiten_traffic_morning[name]["PKW_Total"] for name in present_dirs], dtype=float)
    PKW_afternoon = np.array([PKW_Einheiten_traffic_afternoon[name]["PKW_Total"] for name in present_dirs], dtype=float)
    
    
    #Number of KFZ per side: {"dep_kfz": dep_kfz, "arr_kfz": arr_kfz, "total_kfz": total_kfz}
    side_general = compute_side_sums(flows_present, kfz_general) 
    side_morning = compute_side_sums(flows_present, kfz_morning)
    side_afternoon = compute_side_sums(flows_present, kfz_afternoon)
    
    PKW_side_general = compute_side_sums(flows_present, PKW_general)
    PKW_side_morning = compute_side_sums(flows_present, PKW_morning)
    PKW_side_afternoon = compute_side_sums(flows_present, PKW_afternoon)

    # ---- Per-direction KFZ + Bicycle values for display in Streamlit ----
    per_direction = []
    for name in present_dirs:
        per_direction.append({
            "direction": name,

            "full_day_kfz": float(direction_dic[name]["kfz"]),
            "morning_peak_kfz": float(direction_morning_dic[name]["kfz"]),
            "afternoon_peak_kfz": float(direction_afternoon_dic[name]["kfz"]),

            "full_day_pkw": float(PKW_direction_general_dic[name]["PKW_Total"]),
            "morning_peak_pkw": float(PKW_Einheiten_traffic_morning[name]["PKW_Total"]),
            "afternoon_peak_pkw": float(PKW_Einheiten_traffic_afternoon[name]["PKW_Total"]),

            "full_day_bike": float(direction_dic[name]["rad"]),
            "morning_peak_bike": float(direction_morning_dic[name]["rad"]),
            "afternoon_peak_bike": float(direction_afternoon_dic[name]["rad"]),
        })  
        
        if direction_custom_dic is not None and PKW_Einheiten_traffic_custom is not None:
            per_direction[-1]["custom_kfz"] = float(direction_custom_dic[name]["kfz"])
            per_direction[-1]["custom_pkw"] = float(PKW_Einheiten_traffic_custom[name]["PKW_Total"])
            per_direction[-1]["custom_bike"] = float(direction_custom_dic[name]["rad"]) 
    
    mode = mode.upper().strip()
    use_pkw = mode.startswith("PKW")

    if use_pkw:
        flow_general   = PKW_general
        flow_morning   = PKW_morning
        flow_afternoon = PKW_afternoon

        width_general_sel   = width_PKW_general
        width_morning_sel   = width_PKW_morning
        width_afternoon_sel = width_PKW_afternoon

        unit_label = "PKW_Einheiten"
        suffix_general = "full_day_PKW_Einheiten"
        suffix_morning = "morning_peak_PKW_Einheiten"
        suffix_afternoon = "afternoon_peak_PKW_Einheiten"

        side_general_sel   = PKW_side_general
        side_morning_sel   = PKW_side_morning
        side_afternoon_sel = PKW_side_afternoon
    else:
        flow_general   = kfz_general
        flow_morning   = kfz_morning
        flow_afternoon = kfz_afternoon

        width_general_sel   = width_general
        width_morning_sel   = width_morning_peak
        width_afternoon_sel = width_afternoon_peak

        unit_label = "KFZ"
        suffix_general = "full_day"
        suffix_morning = "morning_peak"
        suffix_afternoon = "afternoon_peak"

        side_general_sel   = side_general
        side_morning_sel   = side_morning
        side_afternoon_sel = side_afternoon
    
    # Generate three plots
    pngs = []
    pngs.append(create_plot(flow_general,   bike_general,   width_general_sel,   flows_present, verkehrszählungsort, suffix_general,   day_start_time,       day_end_time,       side_colors, d_NS, d_WE))
    pngs.append(create_plot(flow_morning,   bike_morning,   width_morning_sel,   flows_present, verkehrszählungsort, suffix_morning,   morning_time_start,   morning_time_end,   side_colors, d_NS, d_WE))
    pngs.append(create_plot(flow_afternoon, bike_afternoon, width_afternoon_sel, flows_present, verkehrszählungsort, suffix_afternoon, afternoon_time_start, afternoon_time_end, side_colors, d_NS, d_WE))

    if direction_custom_dic is not None:
        if use_pkw:
            flow_custom = PKW_custom
            width_custom_sel = width_PKW_custom
            suffix_custom = "custom_1h_PKW_Einheiten"
            side_custom_sel = PKW_side_custom
        else:
            flow_custom = kfz_custom
            width_custom_sel = width_custom
            suffix_custom = "custom_1h"
            side_custom_sel = side_custom

        pngs.append(
            create_plot(
                flow_custom,
                bike_custom,
                width_custom_sel,
                flows_present,
                verkehrszählungsort,
                suffix_custom,
                custom_time_start,
                custom_time_end,
                side_colors,
                d_NS,
                d_WE,
            )
        )

    
    totals = {
    "full_day_kfz": float(np.sum(kfz_general)),
    "morning_peak_kfz": float(np.sum(kfz_morning)),
    "afternoon_peak_kfz": float(np.sum(kfz_afternoon)),

    "full_day_pkw": float(np.sum(PKW_general)),
    "morning_peak_pkw": float(np.sum(PKW_morning)),
    "afternoon_peak_pkw": float(np.sum(PKW_afternoon)),

    "full_day_bike": float(np.sum(bike_general)),
    "morning_peak_bike": float(np.sum(bike_morning)),
    "afternoon_peak_bike": float(np.sum(bike_afternoon)),
    }

    # Add custom totals only if custom exists
    if direction_custom_dic is not None:
        if kfz_custom is not None and bike_custom is not None and PKW_custom is not None:
            totals["custom_kfz"] = float(np.sum(kfz_custom))
            totals["custom_pkw"] = float(np.sum(PKW_custom))
            totals["custom_bike"] = float(np.sum(bike_custom))
    
    meta = {
        "location": verkehrszählungsort,
        "mode": unit_label, 

        "day": {"start": day_start_time, "end": day_end_time},
        "morning_peak": {"start": morning_time_start, "end": morning_time_end},
        "afternoon_peak": {"start": afternoon_time_start, "end": afternoon_time_end},

        "tmin": tmin,
        "tmax": tmax,
        "gamma": gamma,

        # Keep your per_direction as-is OR extend it (see below)
        "per_direction": per_direction,

        # totals now depend on selected mode
        "totals": totals,
    
        "custom": ({"start": custom_time_start, "end": custom_time_end} if direction_custom_dic is not None else None),

        "by_side": {
            "full_day": side_general_sel,
            "morning_peak": side_morning_sel,
            "afternoon_peak": side_afternoon_sel,
            **({"custom": side_custom_sel} if direction_custom_dic is not None else {}),
        },
        
        "sv": {
        "kfz": {
            "full_day": kfz_sv_full,
            "morning_peak": kfz_sv_morning,
            "afternoon_peak": kfz_sv_afternoon,
             **({"custom": kfz_sv_custom} if kfz_sv_custom is not None else {})
        },
        "pkw": {
            "full_day": pkw_sv_full,
            "morning_peak": pkw_sv_morning,
            "afternoon_peak": pkw_sv_afternoon,
            **({"custom": pkw_sv_custom} if pkw_sv_custom is not None else {}),
        },
    },
    }
    
    return pngs, meta